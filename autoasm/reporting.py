"""Reporting (FR10) — query a scan and render a prioritised remediation report.

`scan_summary()` returns structured data used by both the dashboard and the PDF
exporter. `export_pdf()` writes a client-ready report with ReportLab.
"""
from __future__ import annotations

from pathlib import Path

from sqlalchemy import select, func

from .config import DATA_DIR
from .models import (Asset, Correlation, Exposure, Organisation, RiskScore,
                     Scan, get_session)

_SEVERITY_BANDS = [(0.45, "Critical"), (0.30, "High"), (0.15, "Medium"),
                   (0.05, "Low"), (0.0, "Informational")]


def band(score: float) -> str:
    for threshold, label in _SEVERITY_BANDS:
        if score >= threshold:
            return label
    return "Informational"


def scan_summary(scan_id: int) -> dict:
    session = get_session()
    try:
        scan = session.get(Scan, scan_id)
        if scan is None:
            return {}
        org = session.get(Organisation, scan.org_id)
        rows = session.execute(
            select(Exposure, RiskScore, Asset)
            .join(RiskScore, RiskScore.exposure_id == Exposure.id)
            .join(Asset, Asset.id == Exposure.asset_id)
            .where(Asset.org_id == scan.org_id)
            .order_by(RiskScore.composite_score.desc())
        ).all()

        findings = []
        band_counts: dict[str, int] = {}
        for exp, rs, asset in rows:
            cors = session.execute(
                select(Correlation).where(Correlation.exposure_id == exp.id)
            ).scalars().all()
            tags = sorted({c.breach_tag for c in cors if c.breach_tag})
            cves = sorted({c.cve_id for c in cors if c.cve_id})
            kev = any(c.kev_flag for c in cors)
            b = band(rs.composite_score)
            band_counts[b] = band_counts.get(b, 0) + 1
            meta = _CLASS_META.get(exp.cls)
            remediation = (meta[2] if meta else
                           "Review the finding and apply the appropriate control.")
            findings.append({
                "rank": rs.rank, "band": b, "score": rs.composite_score,
                "class": exp.cls, "asset": asset.value, "asset_type": asset.type,
                "description": exp.description, "cvss": exp.cvss_base,
                "criticality": rs.criticality, "severity": rs.severity,
                "breach_relevance": rs.breach_relevance,
                "breach_tags": tags, "cves": cves, "kev": kev,
                "evidence": exp.evidence_ref, "remediation": remediation,
            })
        findings.sort(key=lambda f: f["score"], reverse=True)
        return {
            "scan_id": scan.id,
            "organisation": org.name if org else "",
            "root_domains": org.domain_list() if org else [],
            "started_at": str(scan.started_at),
            "finished_at": str(scan.finished_at),
            "duration_seconds": scan.duration_seconds,
            "asset_count": scan.asset_count,
            "exposure_count": scan.exposure_count,
            "band_counts": band_counts,
            "findings": findings,
        }
    finally:
        session.close()


# --------------------------------------------------------------------------
# Asset inventory — the discovered attack surface itself, independent of
# whether any exposure was found. An EASM tool's first deliverable is "here is
# everything of yours that faces the internet"; findings come second.
# --------------------------------------------------------------------------
def _naive(dt):
    return dt.replace(tzinfo=None) if (dt is not None and dt.tzinfo) else dt


def scan_assets(scan_id: int) -> dict:
    """Return the asset inventory discovered by a scan, grouped and counted.
    Works for a running scan (live) or a completed one, and regardless of
    whether any exposures were found."""
    session = get_session()
    try:
        scan = session.get(Scan, scan_id)
        if scan is None:
            return {}
        org = session.get(Organisation, scan.org_id)
        start, finish = _naive(scan.started_at), _naive(scan.finished_at)

        # Assets are created fresh per scan, so the rows whose first_seen falls in
        # this scan's window are exactly this scan's inventory.
        all_assets = session.execute(
            select(Asset).where(Asset.org_id == scan.org_id)).scalars().all()
        rows = []
        for a in all_assets:
            fs = _naive(a.first_seen)
            if start and fs and fs < start:
                continue
            if finish and fs and fs > finish:
                continue
            rows.append(a)

        ids = [a.id for a in rows]
        ecount: dict[int, int] = {}
        if ids:
            for aid, cnt in session.execute(
                select(Exposure.asset_id, func.count(Exposure.id))
                .where(Exposure.asset_id.in_(ids))
                .group_by(Exposure.asset_id)):
                ecount[aid] = cnt

        items = [{
            "type": a.type, "value": a.value, "source": a.source,
            "criticality": a.criticality, "first_seen": str(_naive(a.first_seen)),
            "exposures": ecount.get(a.id, 0),
        } for a in rows]
        items.sort(key=lambda x: (x["type"], x["value"]))
        breakdown: dict[str, int] = {}
        for it in items:
            breakdown[it["type"]] = breakdown.get(it["type"], 0) + 1
        return {
            "scan_id": scan.id,
            "organisation": org.name if org else "",
            "root_domains": org.domain_list() if org else [],
            "started_at": str(scan.started_at),
            "finished_at": str(scan.finished_at),
            "total": len(items),
            "breakdown": breakdown,
            "assets": items,
        }
    finally:
        session.close()


def export_assets_csv(scan_id: int, out_path: str | None = None) -> str:
    import csv
    data = scan_assets(scan_id)
    if not data:
        raise ValueError(f"No scan with id {scan_id}")
    out = str(out_path or (DATA_DIR / f"AutoASM-NG_Asset_Inventory_scan{scan_id}.csv"))
    with open(out, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["#", "Type", "Asset", "Source", "Criticality", "Exposures",
                    "First seen"])
        for i, a in enumerate(data["assets"], 1):
            w.writerow([i, a["type"], a["value"], a["source"], a["criticality"],
                        a["exposures"], a["first_seen"]])
    return out


def export_assets_xlsx(scan_id: int, out_path: str | None = None) -> str:
    """Arranged asset-inventory workbook: a summary block (counts by type) then a
    sorted, filterable table of every discovered asset."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = scan_assets(scan_id)
    if not data:
        raise ValueError(f"No scan with id {scan_id}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Inventory"
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="D0D7E5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "AutoASM-NG — Asset Inventory"
    ws["A1"].font = title_font
    ws["A2"] = f"Organisation: {data['organisation']}"
    ws["A3"] = f"Scope: {', '.join(data['root_domains'])}"
    ws["A4"] = (f"Scan {data['scan_id']}  ·  {data['total']} assets  ·  "
                f"started {data['started_at']}")

    # summary block (counts by type)
    r = 6
    ws.cell(r, 1, "Summary by type").font = Font(bold=True)
    r += 1
    for c, h in enumerate(["Type", "Count"], 1):
        cell = ws.cell(r, c, h); cell.fill = head_fill; cell.font = head_font
        cell.border = border
    r += 1
    for t in sorted(data["breakdown"]):
        ws.cell(r, 1, t).border = border
        ws.cell(r, 2, data["breakdown"][t]).border = border
        r += 1
    ws.cell(r, 1, "TOTAL").font = Font(bold=True)
    ws.cell(r, 2, data["total"]).font = Font(bold=True)

    # full table
    tbl_start = r + 3
    headers = ["#", "Type", "Asset", "Source", "Criticality", "Exposures",
               "First seen"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(tbl_start, c, h)
        cell.fill = head_fill; cell.font = head_font; cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for i, a in enumerate(data["assets"], 1):
        row = tbl_start + i
        vals = [i, a["type"], a["value"], a["source"], a["criticality"],
                a["exposures"], a["first_seen"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v); cell.border = border

    widths = [5, 12, 60, 22, 11, 11, 22]
    for c, wdt in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = wdt
    ws.freeze_panes = ws.cell(tbl_start + 1, 1)
    last = tbl_start + len(data["assets"])
    ws.auto_filter.ref = f"A{tbl_start}:G{last}"

    out = str(out_path or (DATA_DIR / f"AutoASM-NG_Asset_Inventory_scan{scan_id}.xlsx"))
    wb.save(out)
    return out


# --------------------------------------------------------------------------
# Consultant-style findings XLSX export (uses the workspace findings2de.py reporter)
# --------------------------------------------------------------------------
# Per-class title / asset-type / remediation so the findings report reads like a
# consultant wrote it, not a raw scanner dump.
_CLASS_META = {
    "public-bucket": ("Publicly readable cloud storage bucket", "web",
        "Restrict the bucket policy and ACL to deny public read, enable the "
        "provider Block-Public-Access control, review the contents for sensitive "
        "data, and rotate any exposed material."),
    "exposed-sourcemap": ("First-party JavaScript source map exposed in production", "web",
        "Disable source-map emission in production builds or restrict access to "
        "the .map files, and ensure no secrets are embedded in client-side code."),
    "hardcoded-secret": ("Hardcoded secret in client-delivered code", "web",
        "Remove the secret from client code, rotate it immediately, and load "
        "secrets server-side from a managed secrets store."),
    "exposed-config": ("Sensitive configuration or credential file exposed", "web",
        "Remove the file from the web root, rotate any exposed credentials, and "
        "add deny rules for sensitive paths."),
    "exposed-git": ("Exposed .git repository", "web",
        "Block access to /.git, remove it from the web root, and rotate any "
        "secrets present in the repository history."),
    "exposed-admin-panel": ("Administrative or management endpoint exposed", "web",
        "Restrict the endpoint to trusted networks and authenticated users, and "
        "disable it in production where it is not required."),
    "weak-tls": ("Weak or obsolete TLS configuration", "web",
        "Disable TLS 1.0/1.1 and weak ciphers and enforce TLS 1.2 or higher with "
        "modern cipher suites."),
    "expired-cert": ("Expired TLS certificate", "web",
        "Renew the certificate and automate renewal and expiry monitoring."),
    "subdomain-takeover": ("Subdomain takeover via dangling DNS record", "web",
        "Remove the dangling DNS record or reclaim the referenced backing service."),
    "third-party-hosting": ("Production asset on an unregulated third-party platform", "web",
        "Host production assets on approved and controlled infrastructure and "
        "review the data-handling and contractual controls of any external platform."),
    "exposed-dev-env": ("Non-production environment publicly reachable", "web",
        "Restrict development, staging, and UAT environments to trusted networks "
        "or require authentication, and remove them from public DNS if not needed."),
    "exposed-sensitive-service": ("Sensitive service exposed to the internet", "network",
        "Restrict the service to trusted networks using firewall or security-group "
        "rules, require authentication, and disable it if it need not be public."),
    "exposed-service": ("Service exposed to the internet", "network",
        "Confirm the service must be internet-facing, restrict it by firewall or "
        "security group, and require authentication."),
}


def _host_of(value: str) -> str:
    v = value.split("//")[-1].split("/")[0]
    return v


def export_de_xlsx(scan_id: int, out_path: str | None = None) -> str:
    """Render the scan as a consultant-style findings XLSX report via findings2de.py."""
    import importlib.util
    from .config import FINDINGS2DE_PATH

    data = scan_summary(scan_id)
    if not data:
        raise ValueError(f"No scan with id {scan_id}")
    p = FINDINGS2DE_PATH
    if not p.exists():
        raise FileNotFoundError(f"findings2de.py not found at {p}")
    spec = importlib.util.spec_from_file_location("findings2de", str(p))
    f2de = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(f2de)

    # de-duplicate by (class, asset) and shape into findings2de records
    seen, records = set(), []
    for f in data["findings"]:
        key = (f["class"], f["asset"], f["description"])
        if key in seen:
            continue
        seen.add(key)
        title_base, atype, resolution = _CLASS_META.get(
            f["class"], (f["class"].replace("-", " ").title(), "web",
                         "Review the finding and apply the appropriate control."))
        records.append({
            "title": f"{title_base} ({_host_of(f['asset'])})",
            "severity": f["band"],
            "asset": f["asset"],
            "asset_type": atype,
            "description": f["description"],
            "impact": "",
            "reproduction": "",
            "evidence": f["evidence"],
            "resolution": resolution,
            "status": "VULNERABLE",
            "source_file": "AutoASM-NG",
        })
    out = str(out_path or (DATA_DIR / f"AutoASM-NG_Findings_Report_scan{scan_id}.xlsx"))
    f2de.create_report(records, out, rtype="mixed")
    return out


def export_pdf(scan_id: int, out_path: str | None = None) -> str:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)

    data = scan_summary(scan_id)
    if not data:
        raise ValueError(f"No scan with id {scan_id}")
    out = Path(out_path or (DATA_DIR / f"AutoASM-NG_report_scan{scan_id}.pdf"))

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("Small", parent=styles["Normal"], fontSize=8))
    doc = SimpleDocTemplate(str(out), pagesize=A4, title="AutoASM-NG Report")
    flow = []
    flow.append(Paragraph("AutoASM-NG — External Attack Surface Report",
                          styles["Title"]))
    flow.append(Paragraph(f"Organisation: <b>{data['organisation']}</b>",
                          styles["Normal"]))
    flow.append(Paragraph(f"Scope: {', '.join(data['root_domains'])}",
                          styles["Normal"]))
    flow.append(Paragraph(f"Scan {data['scan_id']} &middot; "
                          f"{data['asset_count']} assets &middot; "
                          f"{data['exposure_count']} exposures &middot; "
                          f"{data['duration_seconds']}s", styles["Small"]))
    flow.append(Spacer(1, 0.4 * cm))

    # severity summary
    flow.append(Paragraph("Exposure summary by risk band", styles["Heading2"]))
    bc = data["band_counts"]
    summary = [["Band", "Count"]] + [[b, str(bc.get(b, 0))] for b in
               ("Critical", "High", "Medium", "Low", "Informational")]
    st = Table(summary, colWidths=[6 * cm, 3 * cm])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    flow.append(st)
    flow.append(Spacer(1, 0.5 * cm))

    # prioritised findings
    flow.append(Paragraph("Prioritised findings", styles["Heading2"]))
    header = ["#", "Band", "Score", "Class", "Asset", "Breach tags / CVEs"]
    table = [header]
    for f in data["findings"][:100]:
        ti = ", ".join(f["breach_tags"] + f["cves"]) + (" [KEV]" if f["kev"] else "")
        table.append([str(f["rank"]), f["band"], f"{f['score']:.3f}", f["class"],
                      Paragraph(f["asset"], styles["Small"]),
                      Paragraph(ti or "-", styles["Small"])])
    ft = Table(table, colWidths=[1*cm, 2*cm, 1.6*cm, 3.2*cm, 5.2*cm, 4*cm],
               repeatRows=1)
    ft.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#eef2f8")]),
    ]))
    flow.append(ft)
    flow.append(Spacer(1, 0.4 * cm))
    flow.append(Paragraph(
        "All checks performed are passive and non-intrusive. Findings are ranked "
        "by RiskScore = asset criticality &times; CVSS severity &times; "
        "breach-relevance (regional breach-pattern weighting).", styles["Small"]))

    doc.build(flow)
    return str(out)
